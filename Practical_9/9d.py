import pandas as pd
import sqlite3 as sq
import re
from openpyxl import load_workbook
import os

# --- GitHub raw CSV that actually exists ---
input_file = "https://raw.githubusercontent.com/plotly/datasets/master/finance-charts-apple.csv"

template_file = r"C:\Users\Administrator\Downloads\Balance-Sheet-Template.xlsx"
output_folder = r"C:\Users\Administrator\Downloads"

# Read CSV
df = pd.read_csv(input_file, encoding="latin-1", engine="python")

# Derive required columns so SQL works
df["Date"] = pd.to_datetime(df["Date"])
df["Year"] = df["Date"].dt.year
df["Quarter"] = df["Date"].dt.quarter
df["Country"] = "USA"
df["Company"] = "Apple"

# Derive some financial‑like fields (proxies)
df["Cash"] = df["AAPL.Close"]
df["Accounts_Receivable"] = df["AAPL.Close"] * 0.1
df["Inventory"] = df["AAPL.Close"] * 0.05
df["Accounts_Payable"] = df["AAPL.Close"] * 0.07
df["Capital_Stock"] = df["AAPL.Close"] * 0.5

# Load into SQLite
conn = sq.connect(":memory:")
df.to_sql("BalanceSheets", conn, if_exists="replace", index=False)

# SQL to get unique reports
sql = """
SELECT 
 Year, Quarter, Country, Company,
 CAST(Year AS INT) || 'Q' || CAST(Quarter AS INT) AS sDate,
 Company || ' (' || Country || ')' AS sCompanyName,
 CAST(Year AS INT) || 'Q' || CAST(Quarter AS INT) || '-' || Company || '-' || Country AS sCompanyFile
FROM BalanceSheets
GROUP BY Year, Quarter, Country, Company
HAVING Year IS NOT NULL;
"""
dates = pd.read_sql_query(re.sub(r"\s\s+", " ", sql), conn)
dates = dates.head(5)

# Fields to fill (only a few, for demo — you can extend)
fields = [
    ("Cash","D16",1), ("Accounts_Receivable","D17",1),
    ("Inventory","D19",1), ("Accounts_Payable","H16",1),
    ("Capital_Stock","H30",1)
]

# Create 5 Excel reports
for _, row in dates.iterrows():
    wb = load_workbook(template_file)
    ws = wb["Balance-Sheet"]

    ws["D3"] = row["sDate"]
    ws["D5"] = row["sCompanyName"]

    for col, cell, sign in fields:
        q = f"""
        SELECT SUM({col}) AS Total
        FROM BalanceSheets
        WHERE Year={int(row['Year'])}
        AND Quarter={int(row['Quarter'])}
        AND Country='{row['Country']}'
        AND Company='{row['Company']}';
        """
        result = pd.read_sql_query(re.sub(r"\s\s+", " ", q), conn)
        ws[cell] = result["Total"][0] * sign

    out = os.path.join(output_folder, f"Report-{row['sCompanyFile']}.xlsx")
    wb.save(out)

print("Done. 5 reports created in Downloads folder.")
